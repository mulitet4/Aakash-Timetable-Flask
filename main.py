# -------
# Imports
# -------
import subprocess
from datetime import datetime
from flask import Flask, render_template, abort, request, send_file
from openpyxl import load_workbook

# ---------
# Constants
# ---------
batch_names: list = [
    'ER-03', 'ER-04', 'FS-01', 'FS-04', 'FS-02',
    'FS-05', 'FS-06', 'FS-08', 'PS', 
    'ER-01', 'ER-02', 'HER-01', 'FS-03', 'FS-07',
    'EW-01'
]
period_codes = [
    'P1', 'P2', 'P3', 'P4', 'P5', 
    'C1', 'C2', 'C3', 'C4', 'C5',
    'M1', 'M2', 'M3', 'M4', 'M5'
    ,'N', 'H', 
]
day_map = {
    "Monday": "EF",
    "Tuesday": "JK",
    "Wednesday": "OP",
    "Thursday": "TU",
    "Friday": "YZ",
    "Saturday": ['AD', 'AE'],
    "Sunday": ['AI', 'AJ'],
}
days = list(day_map.keys())

# --------------
# Initialization
# --------------
batch_map = {}

timetable_wb = load_workbook('./static/src/timetable-aakash.xlsx')
timetable_ws = timetable_wb.active

for batch in batch_names:
    for num in range(1, timetable_ws.max_row+1):
        if batch == timetable_ws['A'+str(num)].value:
            batch_map[batch] = str(num)
            
timetable_wb.close()

app = Flask(__name__)


# ---------
# Rendering
# ---------
@app.route('/', methods=['GET', 'POST'])
def timetable():
    if request.method == 'GET':
        extra_codes = request.args.get('extra_codes')
        if extra_codes:
            extra_codes = extra_codes.split(',')
            extra_codes = [code.strip() for code in extra_codes]
        return render_template(
            'app.html', 
            class_names = batch_names,
            period_codes = period_codes + extra_codes if extra_codes else period_codes,
            days = days
        )
        
    elif request.method == 'POST':
        path, name = prepare_workbook(request.form)
        
        return send_file(path, as_attachment=True, download_name=name)

@app.route('/test', methods=['GET', 'POST'])
def test():
    if request.method == 'GET':
        extra = request.args.get('extra_period_codes')
        print(extra.split(','))
        return "hello"


# ------
# Helper
# ------
def prepare_workbook(form_data):
    date_start = datetime.fromisoformat(form_data.get('date-start')).strftime('%d.%m.%Y')
    date_end = datetime.fromisoformat(form_data.get('date-end')).strftime('%d.%m.%Y')
    
    timetable_wb_src = load_workbook('./static/src/timetable-aakash.xlsx')
    timetable_ws = timetable_wb_src.active
    
    for batch in batch_names:
        for day in days:
            teacher_period_1 = form_data[f"{batch}-{day}-Period-1"]
            teacher_period_2 = form_data[f"{batch}-{day}-Period-2"]
            
            if teacher_period_1 == 'N':
                teacher_period_1 = ''
            if teacher_period_1 == 'H':
                teacher_period_1 = 'Holiday'
                
            if teacher_period_2 == 'N':
                teacher_period_2 = ''
            if teacher_period_2 == 'H':
                teacher_period_2 = 'Holiday'
            
            
            partial_cell_number = batch_map[batch]
            partial_cell_letter_1, partial_cell_letter_2 = day_map[day]
            
            timetable_ws[str(partial_cell_letter_1)+str(partial_cell_number)] = teacher_period_1
            timetable_ws[str(partial_cell_letter_2)+str(partial_cell_number)] = teacher_period_2
    
    path = f"./static/src/timetable-{date_start}-{date_end}.xlsx"
    name = f'timetable-{date_start}-{date_end}.xlsx'
    timetable_wb_src.save(path)
    return path, name

# -------
# Run app
# -------
if __name__ == '__main__':
    subprocess.call([
        'tailwind/tailwindcss-macos-arm64', 
        'build', 
        '-i', 
        'static/src/main.css',
        '-o',
        'static/css/main.css'
        ])
    app.run(debug=True, host="0.0.0.0", port=8000)